"""
Employee data module for the shift scheduling API.

Data Source:
- Development(now): Excel file (xlsx_files/employees_master.xlsx)
- Production(future): Employee data delivered via HTTP request payload
"""
from openpyxl import load_workbook


class Employee:
    """
    Active employee for current month schedule update
    and following month schedule making
    """
    def __init__(
            self,
            employee_id: str,
            first_name: str,
            contract_hours_per_week: float,
            work_calendar_id: str,
            unavailability: list,
            businessunit: str
    ):
        self.employee_id = employee_id
        self.first_name = first_name
        self.contract_hours_per_week = contract_hours_per_week
        self.work_calendar_id = work_calendar_id
        self.unavailability = unavailability
        self.businessunit = businessunit

    @classmethod
    def load_employee_info(cls):
        """Load employees' info from Excel file"""
        workbook = load_workbook(
            filename="./xlsx_files/employees_master.xlsx"
        )
        employee_sheet = workbook["Master"]

        employees = []

        for row in employee_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                employee = cls(
                    employee_id=row[0],
                    first_name=row[1],
                    work_calendar_id=row[2],
                    contract_hours_per_week=row[3],
                    businessunit=row[4],
                    unavailability=[]
                    )
                employees.append(employee)

        unavailability_sheet = workbook["Unavailability"]

        employee_map = {emp.first_name: emp for emp in employees}

        for row in unavailability_sheet.iter_rows(
            min_row=2,
            values_only=True
        ):
            employee_name = row[0]
            unavailability_date = row[1]
            start_time = row[2]
            end_time = row[3]
            if employee_name is None or unavailability_date is None:
                continue
            elif start_time is None and end_time is None:
                unavailable = {
                    "date": unavailability_date.date(),
                    "start_time": None,
                    "end_time": None,
                    "type": "whole_day"}
            else:
                unavailable = {
                    "date": unavailability_date.date(),
                    "start_time": start_time,
                    "end_time": end_time,
                    "type": "partial"}

            if employee_name in employee_map:
                employee_map[employee_name].unavailability.append(
                    unavailable)
            else:
                print(
                    f"No employee found with name '{employee_name}'"
                    )

        return employees


if __name__ == "__main__":
    try:
        # Load all employees
        employees = Employee.load_employee_info()

        print(f"✅ Loaded {len(employees)} employees:")
        print("-" * 40)

        for employee in employees:
            print(f"ID: {employee.employee_id}")
            print(f"Name: {employee.first_name}")
            print(f"Calendar: {employee.work_calendar_id}")
            print(f"Hours: {employee.contract_hours_per_week}")
            print(f"Business Unit: {employee.businessunit}")
            print(f"Unavailability: {employee.unavailability}")
            print("-" * 40)

    except FileNotFoundError:
        print(
            """
Excel file not found! Make sure xlsx_files/
employees_master.xlsx exists
        """)
    except Exception as e:
        print(f"Error: {e}")
